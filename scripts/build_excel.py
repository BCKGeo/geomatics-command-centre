"""
Build the Canadian Open Data Portals Research Excel from jurisdiction + municipality JS data.
Reads the JS files via Node subprocess, then generates a formatted .xlsx.
Run: python scripts/build_excel.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json, subprocess, os, sys

os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Write a temp ESM script to extract data from JS source files
extract_script = os.path.join('scripts', '_extract_data.mjs')
with open(extract_script, 'w', encoding='utf-8') as f:
    f.write("""import { readFileSync } from 'fs';
import { join } from 'path';

const files = ['bc','ab','sk','mb','on','qc','nb','ns','pe','nl','yt','nt','nu','federal'];
const results = {};

for (const f of files) {
  const content = readFileSync(join('src/data/jurisdictions', f + '.js'), 'utf8');
  const idMatch = content.match(/id:"(\\w+)"/);
  const nameMatch = content.match(/name:"([^"]+)"/);
  const abbrMatch = content.match(/abbr:"(\\w+)"/);
  const catRegex = /\\{category:"([^"]+)",links:\\[([^\\]]*?)\\]\\}/g;
  const categories = [];
  let m;
  while ((m = catRegex.exec(content)) !== null) {
    const linkRegex = /\\{n:"([^"]*)",d:"([^"]*)",u:"([^"]*)"(?:,paid:true)?\\}/g;
    const links = [];
    let lm;
    while ((lm = linkRegex.exec(m[2])) !== null) {
      links.push({n: lm[1], d: lm[2], u: lm[3]});
    }
    categories.push({category: m[1], links});
  }
  results[f] = {id: idMatch?.[1]||f, name: nameMatch?.[1]||f, abbr: abbrMatch?.[1]||f.toUpperCase(), categories};
}

const muniContent = readFileSync('src/data/municipalities.js', 'utf8');
const munis = [];
const muniRe = /\\{\\s*name:\\s*"([^"]+)",\\s*province:\\s*"(\\w+)",\\s*lat:\\s*([\\d.-]+),\\s*lon:\\s*([\\d.-]+),\\s*population:\\s*(\\d+),\\s*gisPortal:\\s*(?:"([^"]*?)"|null),\\s*councilUrl:\\s*(?:"([^"]*?)"|null),\\s*surveyStandards:\\s*(?:"([^"]*?)"|null)/g;
let mm;
while ((mm = muniRe.exec(muniContent)) !== null) {
  munis.push({name:mm[1],province:mm[2],lat:+mm[3],lon:+mm[4],population:+mm[5],gisPortal:mm[6]||null,councilUrl:mm[7]||null,surveyStandards:mm[8]||null});
}
results['_municipalities'] = munis;
console.log(JSON.stringify(results));
""")

result = subprocess.run(['node', extract_script], capture_output=True, text=True, shell=False)

os.remove(extract_script)

if result.returncode != 0:
    print("Node error:", result.stderr)
    sys.exit(1)

data = json.loads(result.stdout)
print(f"Loaded {len(data['_municipalities'])} municipalities, {len([k for k in data if k != '_municipalities'])} jurisdictions")

PROV_INFO = {
    'bc': ('British Columbia', '5,400,000'), 'ab': ('Alberta', '4,442,879'),
    'sk': ('Saskatchewan', '1,132,505'), 'mb': ('Manitoba', '1,342,153'),
    'on': ('Ontario', '14,826,276'), 'qc': ('Quebec', '8,604,495'),
    'nb': ('New Brunswick', '789,225'), 'ns': ('Nova Scotia', '969,383'),
    'pe': ('Prince Edward Island', '154,331'), 'nl': ('Newfoundland & Labrador', '510,550'),
    'yt': ('Yukon', '40,232'), 'nt': ('Northwest Territories', '41,070'),
    'nu': ('Nunavut', '36,858'),
}

hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='2F5496')
d_font = Font(name='Arial', size=10)
prov_fill = PatternFill('solid', fgColor='D6E4F0')
tb = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

def get_cat_url(cats, prefix):
    for c in cats:
        if c['category'].lower().startswith(prefix.lower()):
            return c['links'][0]['u'] if c['links'] else None
    return None

def get_cat_all(cats, prefix):
    for c in cats:
        if c['category'].lower().startswith(prefix.lower()):
            return '; '.join([f"{l['n']}: {l['u']}" for l in c['links']])
    return None

wb = openpyxl.Workbook()

# ── Sheet 1: Open Data Portals ──
ws1 = wb.active
ws1.title = 'Open Data Portals'
headers = ['Province/Territory','Entity Name','Entity Type','Population (2021)',
    'Open Data Portal Name','Open Data Portal URL','GIS/Map Viewer Name','GIS/Map Viewer URL',
    'Council Meetings URL','Engineering Standards URL','Professional Bodies',
    'Parcel/Cadastral URL','Base Mapping URL','Imagery URL','LiDAR/Elevation URL',
    'Hydrography URL','Geodetic Control URL','Land Registry URL','Geological Survey URL',
    'Legislation URL','Description / Notes','Last Verified']

for c, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=c, value=h)
    cell.font, cell.fill, cell.border = hdr_font, hdr_fill, tb
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
ws1.freeze_panes = 'A2'
ws1.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

row_num = 2
for pid in ['bc','ab','sk','mb','on','qc','nb','ns','pe','nl','yt','nt','nu']:
    pd = data[pid]
    prov_name, pop = PROV_INFO[pid]
    cats = pd['categories']
    is_terr = pid in ['yt','nt','nu']

    portal_name = viewer_name = None
    for c in cats:
        if c['category'].lower().startswith('open data') and c['links']:
            portal_name = c['links'][0]['n']
        if c['category'].lower().startswith('map viewer') and c['links']:
            viewer_name = c['links'][0]['n']

    row_data = [pd['abbr'], f"{'Territory' if is_terr else 'Province'} of {prov_name}",
        'Territorial Government' if is_terr else 'Provincial Government', pop,
        portal_name, get_cat_url(cats,'Open Data'), viewer_name, get_cat_url(cats,'Map Viewer'),
        None, None, get_cat_all(cats,'Professional'), get_cat_url(cats,'Parcel'),
        get_cat_url(cats,'Base Mapping'), get_cat_url(cats,'Imagery'), get_cat_url(cats,'LiDAR'),
        get_cat_url(cats,'Hydro'), get_cat_url(cats,'Geodetic'), get_cat_url(cats,'Land Registry'),
        get_cat_url(cats,'Geological'), get_cat_url(cats,'Legislation'),
        f"Provincial open data and geospatial resources for {prov_name}.", '2026-04-05']

    for ci, val in enumerate(row_data, 1):
        cell = ws1.cell(row=row_num, column=ci, value=val)
        cell.font, cell.fill, cell.border = d_font, prov_fill, tb
        cell.alignment = Alignment(vertical='top', wrap_text=True)
    row_num += 1

    for m in [x for x in data['_municipalities'] if x['province'] == pd['abbr']]:
        mrow = [pd['abbr'], m['name'], 'Municipal Government', f"{m['population']:,}",
            None, m.get('gisPortal'), None, None, m.get('councilUrl'), m.get('surveyStandards'),
            None,None,None,None,None,None,None,None,None,None,
            f"Municipality in {prov_name}. Pop. {m['population']:,} (2021 Census).", '2026-04-05']
        for ci, val in enumerate(mrow, 1):
            cell = ws1.cell(row=row_num, column=ci, value=val)
            cell.font, cell.border = d_font, tb
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        row_num += 1

for i, w in enumerate([6,35,20,14,25,45,25,45,45,45,60,45,45,45,45,45,45,45,45,45,50,12], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
print(f'Sheet 1: {row_num-2} rows')

# ── Sheet 2: Standards & Reference ──
ws2 = wb.create_sheet('Standards & Reference')
sh = ['Standard/Document','Organization','Scope','URL','Version/Year','Applicable To','Industry','Cost/Access','Description','Last Verified']
for c, h in enumerate(sh, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font, cell.fill, cell.border = hdr_font, hdr_fill, tb
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
ws2.freeze_panes = 'A2'

bc_wb = openpyxl.load_workbook('data/open-data-portals/BC_Open_Data_Portals_Research.xlsx', data_only=True)
sr = 2
for row in bc_wb['Standards & Reference'].iter_rows(min_row=2, max_row=bc_wb['Standards & Reference'].max_row, values_only=True):
    for ci, val in enumerate(row[:len(sh)], 1):
        cell = ws2.cell(row=sr, column=ci, value=val)
        cell.font, cell.border = d_font, tb
        cell.alignment = Alignment(vertical='top', wrap_text=True)
    sr += 1

for row in [
    ['CSA Group','Canadian Standards Association','National','https://www.csagroup.org/','Current','All industries','Standards, Engineering','Varies','Canadian standards development organization.','2026-04-05'],
    ['Standards Council of Canada','Standards Council of Canada','National','https://scc-ccn.ca/','Current','All industries','Standards, International','Free','Canada representative to ISO and IEC.','2026-04-05'],
    ['ISO TC 211 Geographic Information','ISO','International','https://www.isotc211.org/','Current','GIS, Geomatics','Geospatial, Standards','Varies','International standards for geographic information.','2026-04-05'],
    ['CCDC Construction Contracts','CCDC','National','https://www.ccdc.org/','2020/2025','Contractors, consultants','Construction, Engineering','Purchase','Standard form construction contracts. CCDC 2, CCDC 5B.','2026-04-05'],
    ['Transport Canada RPAS Regulations','Transport Canada','National','https://tc.canada.ca/en/aviation/drone-safety','CARs Part IX','RPAS operators','Aviation, RPAS, Surveying','Free','Federal drone regulations. Basic/Advanced certs, SFOC.','2026-04-05'],
    ['NRCan CSRS-PPP Service','Natural Resources Canada','National','https://webapp.csrs-scrs.nrcan-rncan.gc.ca/geod/tools-outils/ppp.php','v5 (2024)','Surveyors','Geodesy, GNSS','Free','Precise Point Positioning. GPS, GLONASS, Galileo.','2026-04-05'],
    ['HRDEM CanElevation Series','Natural Resources Canada','National','https://open.canada.ca/data/en/dataset/957782bf-847c-4644-a757-e383c0057995','Current','Surveyors, GIS','LiDAR, Elevation','Free','Federal high-resolution DEM from LiDAR and satellite.','2026-04-05'],
    ['National Hydro Network (NHN)','Natural Resources Canada','National','https://open.canada.ca/data/en/dataset/a4b190fe-e090-4e6d-881e-b87956c07977','Current','GIS, hydrologists','Hydrography','Free','Federal 1:50K rivers, lakes, and drainage network.','2026-04-05'],
]:
    for ci, val in enumerate(row, 1):
        cell = ws2.cell(row=sr, column=ci, value=val)
        cell.font, cell.border = d_font, tb
        cell.alignment = Alignment(vertical='top', wrap_text=True)
    sr += 1

for i, w in enumerate([35,30,15,55,12,35,28,12,55,12], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
print(f'Sheet 2: {sr-2} standards')

# ── Sheet 3: Metadata & Legend ──
ws3 = wb.create_sheet('Metadata & Legend')
for c, h in enumerate(['Field','Description'], 1):
    cell = ws3.cell(row=1, column=c, value=h)
    cell.font, cell.fill, cell.border = hdr_font, hdr_fill, tb

for ri, (f, d) in enumerate([
    ('Province/Territory','Two-letter code (BC, AB, SK, MB, ON, QC, NB, NS, PE, NL, YT, NT, NU)'),
    ('Entity Name','Official name of municipality or government'), ('Entity Type','Provincial/Territorial/Municipal Government'),
    ('Population (2021)','Statistics Canada 2021 Census'), ('Open Data Portal Name','Portal name'),
    ('Open Data Portal URL','Direct URL'), ('GIS/Map Viewer Name','Viewer name'), ('GIS/Map Viewer URL','Direct URL'),
    ('Council Meetings URL','Agendas, minutes, reports'), ('Engineering Standards URL','Design or subdivision standards'),
    ('Professional Bodies','Semicolon-separated regulatory bodies'), ('Parcel/Cadastral URL','Parcel/cadastral data'),
    ('Base Mapping URL','Topographic base mapping'), ('Imagery URL','Aerial/satellite imagery'),
    ('LiDAR/Elevation URL','LiDAR or DEM data'), ('Hydrography URL','Water/watershed data'),
    ('Geodetic Control URL','Control monuments or GNSS stations'), ('Land Registry URL','Land titles/registry'),
    ('Geological Survey URL','Geological publications/maps'), ('Legislation URL','Professional regulation legislation'),
    ('Description / Notes','Free-text notes'), ('Last Verified','YYYY-MM-DD'),
    ('',''), ('COVERAGE',''),
    ('Provinces covered','10 provinces + 3 territories = 13 jurisdictions'),
    ('Municipalities',f'{len(data["_municipalities"])} major Canadian cities'),
    ('Data source','BCKGeo Dashboard v3 (src/data/jurisdictions/*.js)'),
    ('Maintained by','Ben Koops, AScT - BCKGeo (ben@bckgeo.ca)'), ('Last updated','2026-04-05'),
], 2):
    ws3.cell(row=ri, column=1, value=f).font = Font(name='Arial', bold=True, size=10)
    ws3.cell(row=ri, column=2, value=d).font = d_font
    ws3.cell(row=ri, column=1).border = tb
    ws3.cell(row=ri, column=2).border = tb
    ws3.cell(row=ri, column=2).alignment = Alignment(wrap_text=True)
ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 80

out = 'data/open-data-portals/Canadian_Open_Data_Portals_Research.xlsx'
wb.save(out)
print(f'\nSaved: {out}')
