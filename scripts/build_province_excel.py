"""
Build a BC-style standalone Excel workbook for a single province.

Reads from data/open-data-portals/{prov}_research.json and produces a
formatted .xlsx with 3 sheets matching the BC Open Data Portals Research format.

Usage:
  python scripts/build_province_excel.py --province AB
  python scripts/build_province_excel.py --all
"""

import argparse
import json
import os
import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data" / "open-data-portals"

# Styling constants
HDR_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="2F5496")
DATA_FONT = Font(name="Arial", size=10)
PROV_FILL = PatternFill("solid", fgColor="D6E4F0")
TIER1_FILL = PatternFill("solid", fgColor="E2EFDA")
TIER2_FILL = PatternFill("solid", fgColor="FFF2CC")
TIER3_FILL = None  # No fill for tier 3
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# 27 columns matching the authoritative field mapping
HEADERS = [
    "Province/Territory",
    "Entity Name",
    "Entity Type",
    "Population (2021)",
    "Parent Geography",
    "Open Data Portal Name",
    "Open Data Portal URL",
    "Portal Platform",
    "GIS/Map Viewer Name",
    "GIS/Map Viewer URL",
    "Council Meetings Portal Name",
    "Council Meetings URL",
    "Council Platform",
    "Engineering Standards URL",
    "CAD/Design Standards URL",
    "Construction Standards",
    "Data Formats Available",
    "Coordinate System / Datum",
    "LiDAR/Point Cloud Available",
    "Orthophoto/Imagery Available",
    "WMS/WFS/REST Endpoints",
    "API Endpoint",
    "Data Licence",
    "Contact Department",
    "Industry Focus",
    "Description / Notes",
    "Last Verified",
]

# JSON field -> Excel column index (0-based)
FIELD_MAP = {
    "name": 1,
    "entityType": 2,
    "population": 3,
    "parentGeography": 4,
    "openDataPortalName": 5,
    "openDataPortalUrl": 6,
    "portalPlatform": 7,
    "gisViewerName": 8,
    "gisViewerUrl": 9,
    "councilPortalName": 10,
    "municipalUrl": 11,
    "councilPlatform": 12,
    "engineeringStandardsUrl": 13,
    "cadStandardsUrl": 14,
    "constructionStandards": 15,
    "dataFormats": 16,
    "coordinateSystem": 17,
    "lidarAvailable": 18,
    "orthophotoAvailable": 19,
    "wmsWfsEndpoints": 20,
    "apiEndpoint": 21,
    "dataLicence": 22,
    "contactDepartment": 23,
    "industryFocus": 24,
    "description": 25,
    "lastVerified": 26,
}

COLUMN_WIDTHS = [
    6,   # Province/Territory
    35,  # Entity Name
    22,  # Entity Type
    14,  # Population
    25,  # Parent Geography
    25,  # Open Data Portal Name
    50,  # Open Data Portal URL
    16,  # Portal Platform
    25,  # GIS/Map Viewer Name
    50,  # GIS/Map Viewer URL
    25,  # Council Meetings Portal Name
    50,  # Council Meetings URL
    14,  # Council Platform
    50,  # Engineering Standards URL
    50,  # CAD/Design Standards URL
    20,  # Construction Standards
    30,  # Data Formats Available
    35,  # Coordinate System / Datum
    10,  # LiDAR
    10,  # Orthophoto
    50,  # WMS/WFS/REST
    30,  # API Endpoint
    20,  # Data Licence
    20,  # Contact Department
    20,  # Industry Focus
    50,  # Description
    12,  # Last Verified
]

STANDARDS_HEADERS = [
    "Standard/Document",
    "Organization",
    "Scope",
    "URL",
    "Version/Year",
    "Applicable To",
    "Industry",
    "Cost/Access",
    "Description",
    "Last Verified",
]

STANDARDS_WIDTHS = [35, 30, 15, 55, 12, 35, 28, 12, 55, 12]


def style_header_row(ws, headers):
    """Apply header styling to row 1."""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def write_entity_row(ws, row_num, province_code, entity, is_provincial=False):
    """Write a single entity row to the worksheet."""
    tier = entity.get("tier", 3)

    # Determine row fill
    if is_provincial:
        row_fill = PROV_FILL
    elif tier == 1:
        row_fill = TIER1_FILL
    elif tier == 2:
        row_fill = TIER2_FILL
    else:
        row_fill = TIER3_FILL

    # Column 1: Province code
    cell = ws.cell(row=row_num, column=1, value=province_code)
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    if row_fill:
        cell.fill = row_fill

    # Remaining columns from field map
    for field, col_idx in FIELD_MAP.items():
        value = entity.get(field)
        # Format population with commas
        if field == "population" and isinstance(value, (int, float)) and value:
            value = f"{int(value):,}"
        # Convert booleans to Yes/No
        if field in ("lidarAvailable", "orthophotoAvailable"):
            value = "Yes" if value else "No"
        # Skip None values
        if value is None:
            value = ""

        cell = ws.cell(row=row_num, column=col_idx + 1, value=value)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row_fill:
            cell.fill = row_fill


def build_province_excel(province_code):
    """Build a standalone Excel for one province."""
    province_code = province_code.upper()
    json_path = DATA_DIR / f"{province_code.lower()}_research.json"

    if not json_path.exists():
        print(f"Error: Research JSON not found: {json_path}")
        print(f"Run the research pipeline first: python scripts/research_pipeline.py --province {province_code}")
        return False

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    prov_name = data.get("provinceName", province_code)
    entities = data.get("entities", [])
    standards = data.get("standards", [])

    print(f"\nBuilding Excel for {prov_name} ({province_code})")
    print(f"  Entities: {len(entities)}")
    print(f"  Standards: {len(standards)}")

    wb = openpyxl.Workbook()

    # Sheet 1: Open Data Portals
    ws1 = wb.active
    ws1.title = f"{province_code} Open Data Portals"
    style_header_row(ws1, HEADERS)

    # Sort entities: Tier 1 first, then Tier 2, then Tier 3, alphabetical within each tier
    entities_sorted = sorted(entities, key=lambda e: (e.get("tier", 3), e.get("name", "")))

    row_num = 2
    for entity in entities_sorted:
        write_entity_row(ws1, row_num, province_code, entity)
        row_num += 1

    # Set column widths
    for i, w in enumerate(COLUMN_WIDTHS, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    print(f"  Sheet 1: {row_num - 2} entity rows")

    # Sheet 2: Standards & Reference
    ws2 = wb.create_sheet("Standards & Reference")
    style_header_row(ws2, STANDARDS_HEADERS)

    sr = 2
    for std in standards:
        row_data = [
            std.get("name", ""),
            std.get("organization", ""),
            std.get("scope", ""),
            std.get("url", ""),
            std.get("version", ""),
            std.get("applicableTo", ""),
            std.get("industry", ""),
            std.get("costAccess", ""),
            std.get("description", ""),
            std.get("lastVerified", ""),
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws2.cell(row=sr, column=ci, value=val or "")
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sr += 1

    for i, w in enumerate(STANDARDS_WIDTHS, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    print(f"  Sheet 2: {sr - 2} standards rows")

    # Sheet 3: Metadata & Legend
    ws3 = wb.create_sheet("Metadata & Legend")
    for c, h in enumerate(["Field", "Description"], 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = THIN_BORDER

    meta_rows = [
        ("Province/Territory", "Two-letter province/territory code"),
        ("Entity Name", "Official name of municipality or government entity"),
        ("Entity Type", "City, Town, Village, District Municipality, Rural Municipality, County, etc."),
        ("Population (2021)", "Statistics Canada 2021 Census of Population"),
        ("Parent Geography", "Regional District (BC), County (AB), RM (SK), District (ON), MRC (QC), etc."),
        ("Open Data Portal Name", "Name of the open data portal"),
        ("Open Data Portal URL", "Direct URL to the open data portal"),
        ("Portal Platform", "ArcGIS Hub, CKAN, Socrata, OpenDataSoft, Custom"),
        ("GIS/Map Viewer Name", "Name of web-based GIS/map viewer"),
        ("GIS/Map Viewer URL", "Direct URL to the map viewer"),
        ("Council Meetings Portal Name", "Platform name (CivicWeb, eScribe, Granicus, Custom)"),
        ("Council Meetings URL", "URL for council meeting agendas, minutes, and reports"),
        ("Council Platform", "CivicWeb, eScribe, Granicus, iCompass, Custom"),
        ("Engineering Standards URL", "Design or subdivision engineering standards document"),
        ("CAD/Design Standards URL", "CAD submission standards (Civil 3D, DWG templates)"),
        ("Construction Standards", "MMCD or other construction document standards"),
        ("Data Formats Available", "SHP, GeoJSON, KML, CSV, DWG, LAS/LAZ, GeoTIFF, etc."),
        ("Coordinate System / Datum", "NAD83(CSRS) UTM Zone; Provincial CRS if applicable"),
        ("LiDAR/Point Cloud Available", "Yes/No - whether LiDAR data is available"),
        ("Orthophoto/Imagery Available", "Yes/No - whether orthophoto/imagery is available"),
        ("WMS/WFS/REST Endpoints", "OGC web service endpoint URLs"),
        ("API Endpoint", "REST API type (ArcGIS REST, CKAN API, Socrata SODA, etc.)"),
        ("Data Licence", "OGL, CC-BY, Custom, etc."),
        ("Contact Department", "Department responsible for data (IT/GIS, Planning, Engineering)"),
        ("Industry Focus", "Primary industry relevance (Geospatial, Planning, Engineering)"),
        ("Description / Notes", "Free-text notes about the entity and its data resources"),
        ("Last Verified", "Date the entry was last verified (YYYY-MM-DD)"),
        ("", ""),
        ("COVERAGE", ""),
        ("Province", prov_name),
        ("Total entities", str(len(entities))),
        ("Tier 1 (pop >50k)", str(sum(1 for e in entities if e.get("tier") == 1))),
        ("Tier 2 (pop 10k-50k)", str(sum(1 for e in entities if e.get("tier") == 2))),
        ("Tier 3 (pop <10k)", str(sum(1 for e in entities if e.get("tier") == 3))),
        ("With open data portal", str(sum(1 for e in entities if e.get("openDataPortalUrl")))),
        ("With municipal URL", str(sum(1 for e in entities if e.get("municipalUrl")))),
        ("Data source", "BCKGeo Research Pipeline + Agent Verification"),
        ("Maintained by", "Ben Koops, AScT - BCKGeo (ben@bckgeo.ca)"),
        ("Last updated", data.get("lastUpdated", date.today().isoformat())),
        ("", ""),
        ("TIER LEGEND", ""),
        ("Tier 1 (green)", "Population >50,000. Full 27-field verification."),
        ("Tier 2 (yellow)", "Population 10,000-50,000. Core fields verified."),
        ("Tier 3 (white)", "Population <10,000. Baseline data, portal flagging."),
    ]

    for ri, (field, desc) in enumerate(meta_rows, 2):
        ws3.cell(row=ri, column=1, value=field).font = Font(name="Arial", bold=True, size=10)
        ws3.cell(row=ri, column=2, value=desc).font = DATA_FONT
        ws3.cell(row=ri, column=1).border = THIN_BORDER
        ws3.cell(row=ri, column=2).border = THIN_BORDER
        ws3.cell(row=ri, column=2).alignment = Alignment(wrap_text=True)

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 80

    # Save
    out_path = DATA_DIR / f"{prov_name.replace(' ', '_')}_Open_Data_Portals_Research.xlsx"
    wb.save(out_path)
    print(f"\n  Saved: {out_path}")
    return True


def main():
    parser = argparse.ArgumentParser(description="Build per-province Open Data Portals Excel")
    parser.add_argument("--province", "-p", help="Province code (e.g., AB, ON)")
    parser.add_argument("--all", action="store_true", help="Build Excel for all provinces with research JSON")

    args = parser.parse_args()

    if args.all:
        for json_file in sorted(DATA_DIR.glob("*_research.json")):
            prov_code = json_file.stem.split("_")[0].upper()
            build_province_excel(prov_code)
    elif args.province:
        build_province_excel(args.province)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
